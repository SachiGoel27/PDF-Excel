import pdfplumber
import pandas as pd
from tqdm import tqdm
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side
import re

def extract_tables_(pdf_path):
    combined_data = []
    order_number = 0
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(tqdm(pdf.pages, desc="Processing Pages")):
            try:
                height = page.height
                width = page.width
                cropped_page = page.within_bbox((0, height/2, width, height))
                tables = cropped_page.extract_table(table_settings={
                            "join_tolerance": 7,  
                            "intersection_tolerance": 8,  
                            "horizontal_strategy": "text",
                            "vertical_strategy": "lines_strict",
                            "snap_x_tolerance": 5,
                            "explicit_vertical_lines": [40, 555]
                        })
                
                if tables:
                    order_number=tables[0][5][7:].replace(" ", "")
                    df = pd.DataFrame(tables[3:])
                    if df.shape[1] > 5:
                        df = df.drop(df.columns[[4, 5]], axis=1)
                    
                    df = df[~(df.apply(lambda row: all(cell == "" for cell in row), axis=1))] 
                    df = df[~df[0].astype(str).str.contains("Page", na=False)]

                    data = df.values.tolist()
                    combined_data.extend(data)
            except Exception as e:
                print(f"Error on pathge {page_number+1}: {e}")
    # print(combined_data)
    for index_r, row in enumerate(combined_data):
        # order no.
        combined_data[index_r][1] = combined_data[index_r][1].replace(' ', "")
        # quantity changes
        combined_data[index_r][2] = combined_data[index_r][2].replace('.', "")
        combined_data[index_r][2] = combined_data[index_r][2].replace(',', '.')
        combined_data[index_r][2] = float(combined_data[index_r][2])
        # group #
        if "->" in combined_data[index_r][3]:
            match = re.search(r"\(([^)]+)\)", combined_data[index_r][3])
            extracted_number = match.group(1)[9:].replace(" ", "")
            print(extracted_number)
            combined_data[index_r].append(extracted_number)
        else:
            combined_data[index_r].append(" ")
        # assembly
        combined_data[index_r].append(order_number)
    
    combined_df = ""
    output_stream = ""
    combined_df = pd.DataFrame(combined_data, columns=["Pos", "Order Nr.", "Quantity", "Designation", "Serial from", "Serial to.", "Group #", "Assembly"])
    output_stream = BytesIO()

    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Combined Output')
    
    output_stream.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(output_stream)
    sheet = wb.active

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter 
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  
        sheet.column_dimensions[column_letter].width = adjusted_width

    for cell in sheet[1]: 
        cell.font = Font(bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream