import pdfplumber
import pandas as pd
from tqdm import tqdm
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side


def extract_tables_(pdf_path):
    combined_data = []
    previous_row = None

    with pdfplumber.open(pdf_path) as pdf:
        col_count = 0
        for page_number, page in enumerate(tqdm(pdf.pages, desc="Processing Pages")):
            try:
                page_text = page.extract_text()
                print(page_text.splitlines()[1] + " " + str(page_number))
                if page_text and "Inhaltsverzeichnis" in page_text:
                    continue
                
                elif page_text and "Pos./" in page_text.splitlines()[1]:
                    tables = page.extract_tables(table_settings= {
                        "join_tolerance": 7,  
                        "intersection_tolerance": 8,  
                        "horizontal_strategy": "lines_strict",
                        "snap_x_tolerance": 5,
                        "explicit_vertical_lines": [40, 70, 110, 220, 340, 380, 510, 550]
                    })
                elif page_text and ("ME/" in page_text.splitlines()[1] or "Qty. Unit/" in page_text.splitlines()[1]):
                    tables = page.extract_tables(table_settings={"join_tolerance": 7,  
                                                            "intersection_tolerance": 8,  
                                                            "horizontal_strategy": "lines_strict",
                                                            "snap_x_tolerance": 5,
                                                            "explicit_vertical_lines": [40, 70, 110, 220, 330, 365, 410, 505, 550],
                                                            "explicit_horizontal_lines": [800]
                                                            })
                elif page_text and "Fig./" in page_text.splitlines()[1]:
                    tables = page.extract_tables(table_settings= {
                        "join_tolerance": 7,  
                        "intersection_tolerance": 8,  
                        "horizontal_strategy": "lines_strict",
                        "snap_x_tolerance": 5,
                        "explicit_vertical_lines": [40, 70, 110, 220, 340, 380, 510, 550]
                    })
                elif page_text and "Comment" in page_text.splitlines()[0]:
                    tables = page.extract_tables(table_settings={"join_tolerance": 7,  
                                                            "intersection_tolerance": 8,  
                                                            "horizontal_strategy": "lines_strict",
                                                            "snap_x_tolerance": 5,
                                                            "explicit_vertical_lines": [35, 100, 235, 365, 440, 570],
                                                            "explicit_horizontal_lines": [800]
                                                            })
                elif page_text and "Comment" in page_text.splitlines()[1]:
                    tables = page.extract_tables(table_settings={"join_tolerance": 7,  
                                                            "intersection_tolerance": 8,  
                                                            "horizontal_strategy": "lines_strict",
                                                            "snap_x_tolerance": 5,
                                                            "explicit_vertical_lines": [50, 100, 235, 365, 440, 550],
                                                            "explicit_horizontal_lines": [800]
                                                            })
                else:
                    debug_pic = page.to_image()
                    debug_pic.debug_tablefinder(
                        table_settings={
                            "join_tolerance": 7,  
                            "intersection_tolerance": 8,  
                            "horizontal_strategy": "lines_strict",
                            "snap_x_tolerance": 5,
                            "explicit_vertical_lines": [50, 100, 235, 450, 550],
                            "explicit_horizontal_lines": [800]
                        }
                    )
                    # debug_image_path = f"output_table/debug_page_{page_number+1}.png"
                    # debug_pic.save(debug_image_path)
                    # tables = page.extract_tables(table_settings={"join_tolerance": 7,  
                    #                                         "intersection_tolerance": 8,  
                    #                                         "horizontal_strategy": "lines_strict",
                    #                                         "snap_x_tolerance": 5,
                    #                                         "explicit_vertical_lines": [50, 100, 235, 460, 550],
                    #                                         "explicit_horizontal_lines": [800]
                    #                                         })

                if tables:   
                    for i, table in enumerate(tables): 
                        print(f"Table {i+1}: Columns: {len(table[0])}, Rows: {len(table) - 1}")
                        col_count = len(table[0])                       
                        df = pd.DataFrame(table[1:], columns=table[0])

                        data = df.values.tolist()
                        processed_data = []
                        
                        for row in data:
                            if not row[0].strip() or not row[0].isdigit():
                                if previous_row:
                                    for col_index in range(1, len(row)):
                                        previous_row[col_index] += f" {row[col_index]}".strip()
                                else:
                                    print(f"Warning: Overflow row detected without a previous row: {row}")
                            else:
                                processed_data.append(row)
                                previous_row = row  

                        combined_data.extend(processed_data)

            except Exception as e:
                print(f"Error on pathge {page_number+1}: {e}")
    combined_df = ""
    output_stream = ""
    col_count = len(combined_data[0])
    print(f"First few rows of combined_data: {combined_data[:5]}")
    if col_count == 5 or col_count == 4:
        combined_df = pd.DataFrame(combined_data, columns=["Item no.", "SeboNr", "Description", "Quantity", "Comment"])
        output_stream = BytesIO()
    elif col_count == 7:
        combined_df = pd.DataFrame(combined_data, columns=["Pos./\nFig.", "Ident./\nNo.", "Benennung", "Nomenclature", "Menge/\nQty.", "MPOS~Bemerkung/\nRemark", "s. Seite\nsee Page"])
        output_stream = BytesIO()
    elif col_count == 8:
        combined_df = pd.DataFrame(combined_data, columns=["Fig./\nPos.", "No./\nIdent.", "Nomenclature", "Benennung", "Qty./\nMenge.", "Qty. Unit/", "MPOS~Remark/Bemerkung", "see page\n s. Seite"])
        output_stream = BytesIO()
    # make excel pretty
    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Combined Output')
    
    output_stream.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(output_stream)
    sheet = wb.active

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter 
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  
        sheet.column_dimensions[column_letter].width = adjusted_width

    for cell in sheet[1]: 
        cell.font = Font(bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream
