import pdfplumber
import pandas as pd
from tqdm import tqdm
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side
from typing import List
from collections import defaultdict
from statistics import mean, median


def find_vertical_lines_option1(page, header_fraction=0.25, x_tolerance=20, min_headers=3):
    """
    Infers vertical column boundaries based on the left edge (x0) of bold header text
    and includes the final right boundary (max x1).
    """
    words = page.extract_words(keep_blank_chars=True, use_text_flow=True, extra_attrs=["fontname", "size"])
    header_top = page.height * header_fraction
    header_words = [w for w in words if w['top'] < header_top]

    # Detect bold or large words
    bold_words = [
        w for w in header_words
        if ('Bold' in w.get('fontname', '') or w.get('size', 0) > 7.5)
    ]

    if len(bold_words) < min_headers:
        # Fallback to old logic or return safe defaults
        return [0, 560, round(page.width)]

    # Use x0 of bold words as column starts
    starts = [round(w["x0"]) for w in bold_words]
    ends = [round(w["x1"]) for w in bold_words]
    starts.sort()

    # Cluster close x0s
    clusters = [[starts[0]]]
    for x in starts[1:]:
        if abs(x - clusters[-1][-1]) <= x_tolerance:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    left_edges = [round(mean(c)) for c in clusters]

    # Add final column end (rightmost x1)
    right_most = 560
    verticals = left_edges + [right_most]

    # Cleanup
    verticals = sorted(set(verticals))
    if 0 not in verticals:
        verticals.insert(0, 0)
    if round(page.width) not in verticals:
        verticals.append(round(page.width))

    # Remove near-duplicates
    filtered = [verticals[0]]
    for x in verticals[1:]:
        if x - filtered[-1] > 15:
            filtered.append(x)

    return filtered

import pdfplumber
from statistics import mean


def find_vertical_lines_option2(
    page,
    header_fraction=0.3,
    cluster_tol=4,
    box_merge_tol=3,
    vertical_merge_tol=5,
    x_tolerance=20
):
    """
    Detects vertical lines from header boxes, merging stacked header cells
    and filtering out close duplicates using x_tolerance.

    Args:
        page: pdfplumber page object
        header_fraction: float, portion of page height to treat as header
        cluster_tol: int, clustering tolerance for x positions
        box_merge_tol: int, tolerance for merging boxes with similar x0/x1
        vertical_merge_tol: int, tolerance for stacking boxes vertically
        x_tolerance: int, minimum pixel spacing between final vertical lines

    Returns:
        List of unique, sorted vertical column boundary x-positions.
    """
    header_bottom = page.height * header_fraction
    rects = [r for r in page.rects if r["top"] < header_bottom]

    if not rects:
        return []

    # Group similar boxes together
    merged_boxes = []
    used = [False] * len(rects)

    for i, r1 in enumerate(rects):
        if used[i]:
            continue
        group = [r1]
        for j, r2 in enumerate(rects[i + 1:], start=i + 1):
            if used[j]:
                continue
            same_col = (
                abs(r1["x0"] - r2["x0"]) <= box_merge_tol and
                abs(r1["x1"] - r2["x1"]) <= box_merge_tol
            )
            stacked = (
                abs(r1["top"] - r2["top"]) <= vertical_merge_tol or
                abs(r1["bottom"] - r2["bottom"]) <= vertical_merge_tol
            )
            if same_col and stacked:
                group.append(r2)
                used[j] = True
        used[i] = True
        x0s = [r["x0"] for r in group]
        x1s = [r["x1"] for r in group]
        merged_boxes.append((round(mean(x0s)), round(mean(x1s))))

    # Extract all x0 and x1 values, then cluster
    xs = sorted(set([x for box in merged_boxes for x in box]))
    clustered = [[xs[0]]]
    for x in xs[1:]:
        if abs(x - clustered[-1][-1]) <= cluster_tol:
            clustered[-1].append(x)
        else:
            clustered.append([x])
    verticals = [round(mean(c)) for c in clustered]

    # Final filtering using x_tolerance (optional cleanup for near-duplicates)
    filtered = [verticals[0]]
    for x in verticals[1:]:
        if x - filtered[-1] > x_tolerance:
            filtered.append(x)

    return filtered

def find_vertical_lines_option3(page, min_height=10, tolerance=2):
    x_positions = []

    for line in page.lines:
        x0, x1 = round(line["x0"]), round(line["x1"])
        y0, y1 = line["y0"], line["y1"]
        height = abs(y1 - y0)

        # Only consider vertical lines of sufficient height
        if abs(x0 - x1) < 1 and height >= min_height:
            x_positions.append(x0)

    # Remove near-duplicates by clustering
    x_positions = sorted(set(x_positions))
    clustered = []
    for x in x_positions:
        if not clustered or abs(x - clustered[-1]) > tolerance:
            clustered.append(x)

    return clustered

def flatten_rows(data, add_col=False):
    processed = []
    previous_row = None

    for row in data:
        if "Pos" in str(row[0]).strip():
            continue

        if not add_col and not str(row[0]).strip().isdigit():
            if previous_row:
                for col_index in range(1, len(row)):
                    if previous_row[col_index] is None:
                        previous_row[col_index] = ""
                    if row[col_index]:
                        previous_row[col_index] += f" {row[col_index]}".strip()
            else:
                print(f"Warning: overflow row without a previous one: {row}")
        else:
            if previous_row:
                processed.append(previous_row)
            previous_row = row

    if previous_row:
        processed.append(previous_row)
    return processed


def normalize_headers(df: pd.DataFrame, header_map: dict) -> pd.DataFrame:
    new_columns = {}
    for col in df.columns:
        new_columns[col] = header_map.get(col, col)
    return df.rename(columns=new_columns)

def get_horizontal_lines_from_rects(page, tolerance=1):
    y_positions = []

    for rect in page.rects:
        y0 = round(rect["top"], 1)
        y1 = round(rect["bottom"], 1)
        height = abs(y1 - y0)

        # Treat very short height as horizontal "line" — flat rectangle
        if height <= tolerance:
            y_positions.append(y0)

    # Remove duplicates with tolerance
    y_positions = sorted(set(round(y, 1) for y in y_positions))
    return y_positions


def align_to_standard_schema(df: pd.DataFrame, standard_columns: List[str]) -> pd.DataFrame:
    for col in standard_columns:
        if col not in df.columns:
            df[col] = ""
    return df[standard_columns]


def extract_tables_(pdf_path, option):
    combined_data = []
    STANDARD_COLUMNS = ["Fig./\nPos.", "No./\nIdent.", "Nomenclature", "Benennung", "Qty./\nMenge.", "Qty. Unit/", "MPOS~Remark/Bemerkung", "see page\n s. Seite"]
    HEADER_MAP = {
        "Pos./\nFig.": "Fig./\nPos.",
        "Ident./\nNo.": "No./\nIdent.",
        "Menge\nQty.": "Qty./\nMenge.",
        "MPOS~Bemerkung/Remark" : "MPOS~Remark/Bemerkung",
        "s Seite\nsee page":"see page\n s. Seite",
        "Item no." : "Fig./\nPos",
        "SeboNr": "No./\nIdent.",
        "Fig./\nPos": "Fig./\nPos.",
        "Description" : "Nomenclature",
        "Quantity" : "Qty./\nMenge.",
        "Comment" : "MPOS~Remark/Bemerkung",
        "Ident.\nNo.": "No./\nIdent.",
        "Bemerkung\nRemark" : "MPOS~Remark/Bemerkung",
        "Remark/\nBemerkung" : "MPOS~Remark/Bemerkung",
        "Menge/\nQty.": "Qty./\nMenge.",
        "Item\nno." : "Fig./\nPos",
        "Quanti\nty" : "Qty./\nMenge.",
        "Unit" : "Qty. Unit/",
        "ME/\nQty. Unit" : "Qty. Unit/",
        "siehe S.\nsee page" : "see page\n s. Seite",
        "Qty. Unit/ \nQty. Unit" : "Qty. Unit/",
        "No./\nLevel": "No./\nIdent.",
        "see page\ns. Seite": "see page\n s. Seite",
        "Qty./\nMenge": "Qty./\nMenge.",
        "Iden\nNo" : "No./\nIdent.",
        "MPOS~Bemerkung/\nRemark": "MPOS~Remark/Bemerkung",
        "s. Seite\nsee page": "see page\n s. Seite",
        "Level/\nNo.": "No./\nIdent."

    }
    with pdfplumber.open(pdf_path) as pdf:

        for page_number, page in enumerate(tqdm(pdf.pages, desc="Processing Pages")):
            page_text = page.extract_text()
            try:              
                if page_text and ("Inhaltsverzeichnis" in page_text or "index" in page_text.lower()):
                    continue
                # vertical_line = infer_vertical_lines_from_text(page)
                vertical_line = []
                if option == 1:
                    vertical_line = find_vertical_lines_option1(page)
                elif option == 2:
                    vertical_line = find_vertical_lines_option2(page)
                elif option == 3:
                    vertical_line = find_vertical_lines_option3(page)
                # if not vertical_line:
                #     vertical_line = infer_vertical_lines_from_text(page)
                #     if not vertical_line:
                #         vertical_line = [40, 95, 140, 270, 400, 430, 545]  # final fallback
                if page.lines:
                    tables = page.extract_table(table_settings= {
                        "horizontal_strategy": "lines_strict",
                        "intersection_tolerance": 8,
                        "join_tolerance": 7,
                        "snap_x_tolerance": 5,
                        "explicit_vertical_lines": vertical_line
                    })
                    # debug_pic = page.to_image()
                    # debug_pic.debug_tablefinder(
                    #     table_settings={
                    #     "horizontal_strategy": "lines_strict",
                    #     "intersection_tolerance": 8,
                    #     "join_tolerance": 10,
                    #     "snap_x_tolerance": 10,
                    #     "explicit_vertical_lines": vertical_line
                    #     }
                    # )
                    # debug_image_path = f"output_tables/debug_page_{page_number+1}.png"
                    # debug_pic.save(debug_image_path)
                else:
                    horizontal_line = get_horizontal_lines_from_rects(page)
                    horizontal_line.append(800)
                    tables = page.extract_table(table_settings= {
                        "horizontal_strategy": "lines_strict",
                        "intersection_tolerance": 8,
                        "join_tolerance": 7,
                        "snap_x_tolerance": 5,
                        "explicit_vertical_lines": vertical_line,
                        "explicit_horizontal_lines": horizontal_line
                    })
                    # debug_pic = page.to_image()
                    # debug_pic.debug_tablefinder(
                    #     table_settings={
                    #     "horizontal_strategy": "lines_strict",
                    #     "intersection_tolerance": 8,
                    #     "join_tolerance": 7,
                    #     "snap_x_tolerance": 5,
                    #     "explicit_vertical_lines": vertical_line,
                    #     "explicit_horizontal_lines": horizontal_line
                    #     }
                    # )
                    # debug_image_path = f"output_tables/debug_page_{page_number+1}.png"
                    # debug_pic.save(debug_image_path)
                i = 0
                if tables:   
                    if len(tables) < 2 or not isinstance(tables[0], list):
                        print(f"⚠️ Skipping malformed table on page {page_number + 1}: {tables[0] if tables else 'empty'}")
                        continue  
                    header = [col.strip() if isinstance(col, str) else col for col in tables[0]]
                    # print("Raw headers:", header)
                    raw_rows = tables[1:]
                    df = pd.DataFrame(raw_rows, columns=header)
                    df = normalize_headers(df, HEADER_MAP)
                    # print("Mapped headers:", df.columns.tolist())
                    for col in df.columns:
                        if col not in HEADER_MAP and col not in STANDARD_COLUMNS:
                            print(f"⚠️ Unmapped column: {repr(col)}")
                    flattened_data = flatten_rows(df.values.tolist())
                    df = pd.DataFrame(flattened_data, columns=df.columns)
                    df = align_to_standard_schema(df, STANDARD_COLUMNS)
                    combined_data.append(df)

            except Exception as e:
                print(f"Error on pathge {page_number+1}: {e}")
    if combined_data:
        combined_df = pd.concat(combined_data, ignore_index=True)
        output_stream = BytesIO()
    else:
        combined_df = pd.DataFrame()
        output_stream = BytesIO()
    # combined_df = ""
    # output_stream = ""
    # col_count = len(combined_data[0])
    # if col_count == 5 or col_count == 4:
    #     combined_df = pd.DataFrame(combined_data, columns=["Item no.", "SeboNr", "Description", "Quantity", "Comment"])
    #     output_stream = BytesIO()
    # elif col_count == 6:
    #     combined_df = pd.DataFrame(combined_data, columns=["Pos./\nFig.", "Ident./\nNo.", "Benennung", "Nomenclature", "Menge/\nQty.", "Bemerkung/\nRemark"])
    #     output_stream = BytesIO()
    # elif col_count == 8:
    #     combined_df = pd.DataFrame(combined_data, columns=["Fig./\nPos.", "No./\nIdent.", "Nomenclature", "Benennung", "Qty./\nMenge.", "Qty. Unit/", "MPOS~Remark/Bemerkung", "see page\n s. Seite"])
    #     output_stream = BytesIO()
    # elif odd:
    #     combined_df = pd.DataFrame(combined_data, columns=["Item no.", "SeboNr", "Description", "Quantity", "Unit", "Mpos. No.", "Comment"])
    #     output_stream = BytesIO()
    # else:
    #     combined_df = pd.DataFrame(combined_data, columns=["Pos./\nFig.", "Ident./\nNo.", "Benennung", "Nomenclature", "Menge/\nQty.", "MPOS~Bemerkung/\nRemark", "s. Seite\nsee Page"])
    #     output_stream = BytesIO()
   
    # make excel pretty
    MAX_COL_WIDTH = 50  # set your preferred maximum

    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Combined Output')

    output_stream.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(output_stream)
    sheet = wb.active

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, MAX_COL_WIDTH)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Bold header row
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Apply center alignment + wrap text + thin borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # Save final output
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream